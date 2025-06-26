"""
nuclear_news_indexer.py

Fetches, filters, summarizes, and indexes nuclear-related news articles from
various RSS feeds.

Features:
- Downloads articles from a curated list of science, technology, and policy RSS
  feeds
- Filters articles by nuclear-related keywords
- Summarizes and translates content using Azure OpenAI
- Uploads processed articles to Azure Cognitive Search
- Logs uploads and saves results to an Excel file

Requirements:
- Azure Key Vault for secret management
- Azure OpenAI and Azure Cognitive Search resources
- Python packages: feedparser, requests, openpyxl, azure-identity, \
  azure-keyvault-secrets, azure-search-documents, openai, python-dotenv

Usage:
- Configure environment variables for Azure resources and Key Vault
- Run the script to fetch and process news articles from the past week
"""

# Version 2.8.2: Enhanced HTTP headers to better emulate real browsers

import json
import logging
import os
import sys
import uuid
from datetime import datetime, timedelta, timezone
from typing import Any, Optional

import feedparser
import requests
from azure.core.credentials import AzureKeyCredential
from azure.identity import DefaultAzureCredential
from azure.keyvault.secrets import SecretClient
from azure.search.documents import SearchClient
from dotenv import load_dotenv
from openai import AzureOpenAI
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

load_dotenv()

# Configure logging
LOG_DIR = os.path.join(os.path.dirname(__file__), "logs")
LOG_FILE = os.path.join(
    LOG_DIR, f'news_indexer_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'
)
os.makedirs(LOG_DIR, exist_ok=True)
logging.basicConfig(
    filename=LOG_FILE,
    filemode="a",
    format="%(asctime)s [%(levelname)s] %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)
# Add console handler
console_handler = logging.StreamHandler(sys.stdout)
console_handler.setLevel(logging.INFO)
console_handler.setFormatter(
    logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
)
logger.addHandler(console_handler)

feeds = [
    # News & Science
    "https://rss.nytimes.com/services/xml/rss/nyt/Science.xml",
    "https://www.sciencedaily.com/rss/matter_energy/nuclear_energy.xml",
    "https://www.sciencedaily.com/rss/matter_energy/quantum_physics.xml",
    "https://www.nasa.gov/rss/dyn/breaking_news.rss",
    "https://phys.org/rss-feed/breaking/",
    "https://www.sciencedaily.com/rss/all.xml",
    "https://www.newscientist.com/feed/home",
    "https://www.science.org/rss/news_current.xml",
    "https://www.the-scientist.com/rss",
    "https://techcrunch.com/feed/",
    "https://www.theverge.com/rss/index.xml",
    "https://www.wired.com/feed/rss",
    "https://feeds.arstechnica.com/arstechnica/index",
    "https://www.engadget.com/rss.xml",
    # Global News
    "https://www.aljazeera.com/xml/rss/all.xml",
    "https://timesofindia.indiatimes.com/rssfeeds/4719148.cms",
    "https://www3.nhk.or.jp/rss/news/cat0.xml",
    "https://www.cbc.ca/cmlink/rss-world",
    "https://canarymedia.com/rss.rss",
    "https://apnews.com/index.rss",
    "https://rss.nytimes.com/services/xml/rss/nyt/US.xml",
    "https://rss.nytimes.com/services/xml/rss/nyt/World.xml",
    # Nuclear Policy, Regulation, and Advocacy
    "https://nuclear-news.net/feed/",
    "https://neutronbytes.com/feed/",
    "https://www.iaea.org/rss/news.xml",
    "https://thebulletin.org/search-feed",
    "https://carnegieendowment.org/feed/proliferation-news",
]

keywords = [
    # Core nuclear terms
    "nuclear",
    "LPO",
    "DOE",
    "NRC",
    "IAEA",
    "tritium",
    "uranium",
    "atomic",
    "fusion",
    "fission",
    "reactor",
    "plasma",
    "neutron",
    "isotope",
    "radiation",
    "particle",
    "quantum",
    # Advanced technical terms
    "deuterium",
    "tokamak",
    "breeder reactor",
    "high energy physics",
    "radioisotope",
    "criticality",
    "chain reaction",
    "nuclear waste",
    "spent fuel",
    "containment",
    "reprocessing",
    "coolant leak",
    "thermal neutron",
    "cross-section",
    "nuclear fuel cycle",
    # Policy & disarmament
    "arms control",
    "nuclear treaty",
    "non-proliferation",
    "START",
    "CTBT",
    "NPT",
    "deterrence",
    "disarmament",
    # Emerging science terms
    "muon",
    "stellarator",
    "quark",
    "superconducting",
    "fusion ignition",
    "neutrino",
    "synchrotron",
]


def fetch_feed_with_timeout(
    url: str, timeout: int = 10
) -> Optional[feedparser.FeedParserDict]:
    """
    Fetches and parses an RSS/Atom feed from the given URL with a timeout and
    browser-like headers.

    Args:
        url (str): The URL of the RSS/Atom feed.
        timeout (int, optional): Timeout in seconds for the HTTP request.
            Defaults to 10.

    Returns:
        feedparser.FeedParserDict or None: Parsed feed object, or None if
            fetch fails.
    """
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept": (
            "text/html,application/xhtml+xml,application/xml;"
            "q=0.9,*/*;q=0.8"
        ),
        "Accept-Encoding": "gzip, deflate",
        "Accept-Language": "en-US,en;q=0.5",
        "Connection": "keep-alive",
        "Referer": "https://www.google.com/",
        "DNT": "1",
    }
    try:
        resp = requests.get(url, timeout=timeout, headers=headers)
        resp.raise_for_status()
        logger.info(f"Fetched feed: {url}")
        return feedparser.parse(resp.content)
    except Exception as e:
        logger.warning(f"Failed to fetch {url}  ̲ {e}")
        return None


def matches_keywords(text: str) -> bool:
    """
    Checks if the provided text contains any of the defined nuclear-related
    keywords.

    Args:
        text (str): The text to search for keywords.

    Returns:
        bool: True if any keyword is found, False otherwise.
    """
    text = text.lower()
    return any(k in text for k in keywords)


def process_feed(
    url: str,
    one_week_ago: datetime,
    existing_urls: set,
    ws,
    keywords: list,
    client,
    model_name: str,
    search_client,
    logger,
) -> None:
    """
    Process a single RSS feed: fetch, filter, summarize, upload, and log
    articles.
    """
    logger.info(f"Parsing feed: {url}")
    feed = fetch_feed_with_timeout(url)
    if not feed:
        return
    logger.info(f"Found {len(feed.entries)} entries.")
    for entry in feed.entries:
        process_entry(
            entry,
            feed,
            one_week_ago,
            existing_urls,
            ws,
            keywords,
            client,
            model_name,
            search_client,
            logger,
        )


def is_entry_recent(entry, one_week_ago: datetime, logger) -> bool:
    """
    Returns True if the entry is recent (published within the last week),
    else False.
    Logs and handles parsing errors gracefully.
    """
    try:
        published_dt = extract_published_dt(entry)
    except Exception as e:
        title = getattr(entry, "title", entry.get("title", ""))
        logger.warning(
            "Failed to parse published date for entry: %s",
            title,
        )
        logger.warning(
            "error: %s",
            e,
        )
        published_dt = datetime.now(timezone.utc)
    if published_dt < one_week_ago:
        msg = "Skipping old article: %s" % getattr(
            entry, "title", entry.get("title", "")
        )
        logger.info(msg)
        return False
    return True


def is_entry_duplicate(entry, existing_urls: set, logger) -> bool:
    if entry.link in existing_urls:
        logger.info(f"Skipping duplicate URL: {entry.title}")
        return True
    return False


def get_entry_summary(entry, client, model_name: str, logger) -> str:
    content = entry.get("summary", "")
    if not content:
        logger.warning("No summary available in RSS feed.")
        return ""
    try:
        translation_prompt = (
            "Translate this to English (if not already), then summarize:\n"
            f"{content[:4000]}"
        )
        logger.info(
            "Sending translation + summary request to OpenAI for: %s",
            entry.title,
        )
        response = client.chat.completions.create(
            model=model_name,
            messages=[{"role": "user", "content": translation_prompt}],
            temperature=0.3,
        )
        summary = response.choices[0].message.content.strip()
        logger.info(f"Got summary for: {entry.title}")
        return summary
    except Exception as e:
        logger.error(f"Error summarizing article: {e}")
        return ""


def upload_entry_to_search(doc: dict, search_client, logger) -> bool:
    try:
        result = search_client.upload_documents(documents=[doc])
        status = (
            result[0].status_code
            if hasattr(result[0], "status_code")
            else "Success"
        )
        logger.info("Uploaded: %s Status: %s", doc["title"], status)
        return True
    except Exception as e:
        logger.error(f"Error uploading to Azure Search: {e}")
        return False


def extract_published_dt(entry) -> datetime:
    """
    Extracts and returns a timezone-aware published datetime from an RSS entry.
    Falls back to current UTC time if parsing fails.
    """
    published_str = entry.get("published", None)
    if not published_str:
        return datetime.now(timezone.utc)
    published_parsed = getattr(entry, "published_parsed", None)
    try:
        if not published_parsed:
            return datetime.now(timezone.utc)
        if isinstance(published_parsed, datetime):
            published_dt = published_parsed
        elif isinstance(published_parsed, tuple):
            if len(published_parsed) == 6:
                published_dt = datetime(*published_parsed, tzinfo=timezone.utc)
            elif len(published_parsed) > 6:
                published_dt = datetime(*published_parsed[:6])
            else:
                published_dt = datetime.now(timezone.utc)
        else:
            published_dt = datetime.now(timezone.utc)
        # Ensure timezone-aware (UTC)
        if (
            published_dt.tzinfo is None
            or published_dt.tzinfo.utcoffset(published_dt) is None
        ):
            published_dt = published_dt.replace(tzinfo=timezone.utc)
    except (TypeError, ValueError):
        published_dt = datetime.now(timezone.utc)
    return published_dt


def process_entry(
    entry,
    feed,
    one_week_ago: datetime,
    existing_urls: set,
    ws,
    keywords: list,
    client,
    model_name: str,
    search_client,
    logger,
) -> None:
    """
    Process a single feed entry: filter, summarize, upload, and log.
    """
    if not is_entry_recent(entry, one_week_ago, logger):
        return
    combined_text = entry.title + " " + entry.get("summary", "")
    if not matches_keywords(combined_text):
        logger.info(f"Skipping (no keyword match): {entry.title}")
        return
    if is_entry_duplicate(entry, existing_urls, logger):
        return
    summary = get_entry_summary(entry, client, model_name, logger)
    if not summary:
        return
    content = entry.get("summary", "")
    published_dt = extract_published_dt(entry)
    doc = {
        "id": str(uuid.uuid4()),
        "title": entry.title,
        "summary": summary,
        "url": entry.link,
        "author": entry.get("author", "Unknown"),
        "tags": [k for k in keywords if k in content.lower()],
        "publishedDate": published_dt.isoformat(),
        "source": feed.feed.get("title", "RSS Source"),
        "content": content[:8000],
    }
    if not upload_entry_to_search(doc, search_client, logger):
        return
    if ws is None:
        raise RuntimeError("Worksheet is None")
    worksheet = get_worksheet(ws)
    worksheet.append(
        [
            doc["title"],
            summary,
            doc["url"],
            doc["author"],
            ", ".join(doc["tags"]),
            doc["publishedDate"],
            doc["source"],
        ]
    )
    existing_urls.add(doc["url"])
    with open("upload.log", "a", encoding="utf-8") as logf:
        logf.write(json.dumps(doc, indent=2) + "\n\n")


def get_azure_clients_and_secrets():
    """
    Lazily fetch Azure secrets and instantiate clients.
    Call only inside main().
    Returns:
        client (AzureOpenAI): OpenAI client
        model_name (str): OpenAI deployment name
        search_client (SearchClient): Azure Search client
    """
    key_vault_url = os.getenv("KEY_VAULT_URL")
    credential = DefaultAzureCredential()
    secret_client = SecretClient(
        vault_url=key_vault_url, credential=credential
    )
    openai_api_key = secret_client.get_secret("AI-OPENAI-KEY").value
    openai_api_base = secret_client.get_secret("AI-OPENAI-ENDPOINT").value
    openai_deployment = secret_client.get_secret("AI-OPENAI-DEPLOYMENT").value
    search_api_key = secret_client.get_secret("AI-SEARCH-PRIMARY-KEY").value
    search_api_endpoint = secret_client.get_secret("AI-SEARCH-ENDPOINT").value
    client = AzureOpenAI(
        api_key=openai_api_key,
        api_version="2024-12-01-preview",
        azure_endpoint=openai_api_base,
    )
    model_name = openai_deployment
    search_client = SearchClient(
        endpoint=search_api_endpoint,
        index_name="news-articles-index",
        credential=AzureKeyCredential(search_api_key),
    )
    return client, model_name, search_client


def get_worksheet(ws: Any) -> Worksheet:
    if ws is None or not isinstance(ws, Worksheet):
        raise RuntimeError("Worksheet is None or not a Worksheet instance")
    return ws


def main() -> None:
    """
    Main execution function for fetching, filtering, summarizing, and indexing
    nuclear-related news articles.
    Handles feed parsing, keyword filtering, summarization, Azure Search
    upload, and Excel logging.
    """
    client, model_name, search_client = get_azure_clients_and_secrets()
    output_dir = os.path.join(os.path.dirname(__file__), "output")
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_file = os.path.join(output_dir, f"news_results_{timestamp}.xlsx")
    headers = [
        "Title",
        "Summary",
        "URL",
        "Author",
        "Tags",
        "PublishedDate",
        "Source",
    ]
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(excel_file)
    logger.info(f"Excel output saved to: {os.path.abspath(excel_file)}")
    wb = load_workbook(excel_file)
    worksheet = get_worksheet(wb.active)
    existing_urls = {
        row[2]
        for row in worksheet.iter_rows(min_row=2, values_only=True)
        if row[2]
    }
    one_week_ago = datetime.now(timezone.utc) - timedelta(days=7)
    for url in feeds:
        process_feed(
            url,
            one_week_ago,
            existing_urls,
            worksheet,
            keywords,
            client,
            model_name,
            search_client,
            logger,
        )
    wb.save(excel_file)
    logger.info("Job complete.")
    logger.info(
        "Excel output saved to: %s",
        os.path.abspath(excel_file),
    )


if __name__ == "__main__":
    main()
